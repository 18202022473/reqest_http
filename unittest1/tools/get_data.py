from lemon.unittest1.tools.project_path import *
from openpyxl import load_workbook


class GetData:
    token = None
    phone = load_workbook(excel_path)['data3'].cell(1, 1).value
    password = 'af8f9dffa5d420fbc249141645b962'
# wb = load_workbook(excel_path)
# sheet = wb['data3']
# sheet.cell(1, 1).value
