from openpyxl import load_workbook
from lemon.unittest1.tools.config import Config
from lemon.unittest1.tools.project_path import *
from lemon.unittest1.tools.get_data import GetData


def get_excel(excel_file):
    title = []
    user_data = []
    '''获取config中表单名和对应表单要执行的用例的字典'''
    config_data = eval(Config().config(config_path, 'MODE', 'mode'))
    '''遍历字典获取表单名excel_sheet，和要执行的用例mode'''
    for excel_sheet in config_data:
        wb = load_workbook(excel_file)
        sheet = wb[excel_sheet]
        mode = config_data[excel_sheet]
        '''获取Excel中的标题行'''
        for i in range(1, sheet.max_column + 1):
            res = sheet.cell(1, i).value
            title.append(res)
        '''遍历单个表单中的所有数据，以列表嵌套字典的形式'''
        for j in range(2, sheet.max_row + 1):
            data_cell = {}
            for i in range(1, sheet.max_column + 1):
                '''如果存在'$',就将它替换成phone'''
                if title[i - 1] == 'requests_data':
                    if sheet.cell(j, i).value.find('${}') != -1:
                        data_cell[title[i - 1]] = sheet.cell(j, i).value.replace('${}', str(getattr(GetData, 'phone')))
                    elif sheet.cell(j, i).value.find('${1}') != -1:
                        data_cell[title[i - 1]] = sheet.cell(j, i).value.replace('${1}', str(getattr(GetData, 'phone')+1))
                else:
                    data_cell[title[i - 1]] = sheet.cell(j, i).value
            if mode == 'all' or j - 1 in mode:
                user_data.append(data_cell)
            else:
                pass
    phone = getattr(GetData, 'phone')+2
    do_excel(excel_path, 'data3', 0, phone, 1)
    return user_data


def do_excel(excel_file, excel_sheet, case_id, value, write_column):
    wb = load_workbook(excel_file)
    sheet = wb[excel_sheet]
    sheet.cell(case_id + 1, write_column).value = value
    wb.save(excel_file)


if __name__ == '__main__':
    print(get_excel(excel_path))
